"""Minimal .xlsx parser — sheets + rows as structured text."""

from pathlib import Path

from harness.ingestion.models import ParsedDoc


def parse_xlsx(path: Path, tenant_id: str | None = None) -> ParsedDoc:
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX parsing. "
            "Install: pip install openpyxl"
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n## Sheet: {sheet_name}")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [str(c) if c is not None else "" for c in row]
            line = " | ".join(vals).strip()
            if line:
                lines.append(f"Row {row_idx}: {line}")
    wb.close()
    return ParsedDoc(
        source_filename=path.name,
        content_type="xlsx",
        text="\n".join(lines),
        metadata={"sheets": wb.sheetnames},
        tenant_id=tenant_id,
    )
